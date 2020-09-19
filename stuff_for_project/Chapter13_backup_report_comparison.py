# Compare SQL backup reports
# 1. Get excel file names. name format is: customer_sql_report.xlsx
# 2. ask to enter customer name, if none provided and there are more than 2 .xlsx files in folder - exit, compare modification time os.path.getmtime(path)
# 3. read file 1. Create data structure: {server: database, True}
# 4. read file 2
# 5. go through 1st dict, compare to 2nd one, modify it accordingly
# 6. create and excel WB, write header, write 1st dict
# 7. save excel file, report on completion
# 8. move old file to ../old_files/
# P.S. pip install --user -U openpyxl==2.6.2


import re, openpyxl, sys, pathlib, pyinputplus, pprint, logging, os
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string, get_column_letter
logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s -%(levelname)s -  %(message)s')


def get_names():
    """
    Looks for *sql_report* files in cwd, tries to find exactly 2 for same customer if it can determine customer, otherwise
    asks for customer name or fails
    :return: two .xlsx file paths that will be used later in data comparison
    """
    number_of_reports_in_cwd = len(list(pathlib.Path.cwd().glob("*sql_report*.xlsx")))  # get number of reports that match pattern
    if number_of_reports_in_cwd > 2:
        print("There are more than 2 *sql_report*.xlsx files in current directory - you need to clarify customer")
        try:
            customer_name = pyinputplus.inputStr(prompt="Please enter customer name", blockRegexes=["[0-9+=*/$%^#@!]"], limit=3)  # get customer name
            if len(list(pathlib.Path.cwd().glob(f"{customer_name}_*sql_report*.xlsx"))) != 2:  # more / less than 2 reports in cwd for chosen customer
                print("This script require 2 *sql_report.xlsx files to be put in current working directory per customer. Exiting")
                sys.exit("Not enough .xlsx files in folder for selected customer")
        except pyinputplus.RetryLimitException:
            print("You have failed to enter valid customer name [a-zA-Z-], exiting")
            sys.exit("Invalid customer name")
    elif number_of_reports_in_cwd == 2:  # do not need customer name, we EXPECT that both are for same customer
        print("There are 2 *sql_report.xlsx files in folder. Modification date will be used to determine the old one")
        customer_name = re.search("([a-zA-Z-]+_)(.*sql_report.*)", str(list(pathlib.Path.cwd().glob("*sql_report*.xlsx"))[0].name)).group(1).rstrip("_")  # extract name
        if len(list(pathlib.Path.cwd().glob(f"{customer_name}_*sql_report*.xlsx"))) != 2:  # just in case if second file is for another customer
            print("This script require 2 *sql_report.xlsx files to be put in current working directory per customer. Exiting")
    else:
        print("This script require at least 2 *sql_report.xlsx files to be put in current working directory. Exiting")
        sys.exit("Not enough .xlsx files in folder")
    logging.info(f"customer_name detected is: {customer_name}, total number of report files in cwd: {number_of_reports_in_cwd}\n Report files that will be used are {list(pathlib.Path.cwd().glob(f'{customer_name}_*sql_report*.xlsx'))}")
    return list(pathlib.Path.cwd().glob(f"{customer_name}_*sql_report*.xlsx"))


def prepare_data(path):
    """
    reads excel file names provided by get_names()
    :param path: path to excel file to process
    :return: dictionary where keys are servers and values are
    dictionaries with db_names as keys and backup status as values
    """
    logging.info(f"Processing SQL backup report - {path}")
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        logging.critical(f"Critical error! Workbook cannot be opened! Exiting! \nError message: {e}")
        sys.exit(1)
    sheet = wb.active
    data = dict()
    for i in range(2, sheet.max_row + 1):
        server = sheet.cell(row=i, column=1).value
        db_name = str(sheet.cell(row=i, column=2).value).replace("ø", "o").replace("Ø", "O")
        in_backup = sheet.cell(row=i, column=3).value
    # Data_model = {"server":
    #                     {
    #                     "db_name": in_backup,
    #                     "db_name2": in_backup
    #                     }
    #              }
        data.setdefault(server, {})  # we set dictionary values so we wont fuck up with missing server key
        data[server].setdefault(db_name, in_backup)  # we put data for current server's DB backup status into dict. value, we EXPECT that DB names are unique
    return data

## Testing part
# logging.warning("Changing directory")
# os.chdir(r"C:\Users\yaizk\OneDrive\Desktop\to_delete")

file_old, file_new = sorted(get_names(), key=lambda x: os.path.getmtime(x))
logging.info(f"Old report detected as {file_old.name}, New report detected as {file_new.name}")
data_old = prepare_data(file_old)
data_new = prepare_data(file_new)

for server in data_new.keys():  # we go through data returned by prepare_data() and make changes to data_new accordingly
    for database in data_new[server]:
        data_old.setdefault(server, {})  # in case if new server appeared in new report that not present in old report
        data_old[server].setdefault(database, data_new[server][database])
        if database not in data_old[server]:
            data_new[server][database] += " New_DB"
        elif database in data_old[server] and "Yes" in data_old[server][database] and "Yes" in data_new[server][database]:
            data_new[server][database] = "Yes"
        elif database in data_old[server] and "Yes" in data_old[server][database] and "No" in data_new[server][database]:
            data_new[server][database] += " Backup_disabled"
        elif database in data_old[server] and "No" in data_old[server][database] and "Yes" in data_new[server][database]:
            data_new[server][database] += " Backup_enabled"
        elif database in data_old[server] and "No" in data_old[server][database] and "No" in data_new[server][database]:
            data_new[server][database] = "No"

result_wb = openpyxl.Workbook()
r_sheet = result_wb.active
r_sheet.title = "SQL_backup_report_with_changes"  # making sheet beautiful
header_style = Font(bold=True)
r_sheet.cell(row=1, column=1).value = "Server"
r_sheet.cell(row=1, column=1).font = header_style
r_sheet.cell(row=1, column=2).value = "Database"
r_sheet.cell(row=1, column=2).font = header_style
r_sheet.cell(row=1, column=3).value = "In Backup"
r_sheet.cell(row=1, column=3).font = header_style
r_sheet.cell(row=1, column=4).value = "Backup Changes"
r_sheet.cell(row=1, column=4).font = header_style
r_sheet.freeze_panes = "A2"

row = 2
for server in data_new.keys():
    for database in data_new[server]:
        r_sheet.cell(row=row, column=1).value = server
        r_sheet.cell(row=row, column=2).value = database
        try:
            r_sheet.cell(row=row, column=3).value = str(data_new[server][database]).split()[0]  # we do not know if there are 2 words or one in this cell
        except:
            logging.error(f"Error ! server = {server}, database = {database}")
        try:
            r_sheet.cell(row=row, column=4).value = str(data_new[server][database]).split()[1]  # so we have to use try to not to get errors
        except IndexError:
            pass
        row += 1

try:
    result_wb.save("result.xlsx")
except PermissionError:
    logging.error("Cannot save resulting workbook, trying to use reserve name - debug_result.xlsx")
    result_wb.save("debug_result.xlsx")

